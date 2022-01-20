"""
File containing Excel processing functionality.
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from models import Results


def format_results(searh_id):
    """
    This method populates the template excel file, using openpyxl, with information from the
    database for a given search and saves it within the program repository.
    """
    # open the BOM Excel template
    spreadsheet = load_workbook('BOMOutputTemplate/BOMOutput.xlsx')

    # Retrieve the dataset which will be used to populate the template
    results_data = Results.query.filter(Results.searchnumber == searh_id).all()

    # Create a write object
    write = spreadsheet.active

    # Variable which increments the writer down cells.
    increment_line = 0

    # Running total of price
    totalprice = 0

    for result in results_data:

        # If the part has not been found, then use this row template
        if result.stock <= result.stockrequired or Results.stock == 0:

            write[f'C{str(8 + increment_line)}'] = result.partnumber
            write[f'D{str(8 + increment_line)}'] = result.supplier
            write[f'E{str(8 + increment_line)}'] = "Not Found"
            write[f'F{str(8 + increment_line)}'] = "Not Found"
            write[f'G{str(8 + increment_line)}'] = "Not Found"
            write[f'H{str(8 + increment_line)}'] = "Not Found"
            write[f'I{str(8 + increment_line)}'] = "Not Found"

            # Colour code the 'stock' column red
            write[f'E{str(8 + increment_line)}'].fill = \
                PatternFill(patternType='solid', fgColor='00800000')

        # If the part has  been found, then use this row template
        else:
            write[f'C{str(8 + increment_line)}'] = result.partnumber
            write[f'D{str(8 + increment_line)}'] = result.supplier
            write[f'E{str(8 + increment_line)}'] = result.stock
            write[f'F{str(8 + increment_line)}'] = result.stockrequired
            write[f'G{str(8 + increment_line)}'] = result.priceperunit
            write[f'H{str(8 + increment_line)}'] = result.totalprice
            write[f'I{str(8 + increment_line)}'] = result.link

            # Colour code the 'stock' column green
            write[f'E{str(8 + increment_line)}'].fill = \
                PatternFill(patternType='solid', fgColor='00008000')

            # Increment total price accordingly
            totalprice += result.totalprice

        increment_line += 1

    # Populate total price cell
    write['H5'] = f'   Total price: £{round(totalprice, 2)}'
    write['H5'].font = Font(size=19)

    # Save and close file
    spreadsheet.save(f'SearchResults/BOMSearch{searh_id}results.xlsx')
    spreadsheet.close()
