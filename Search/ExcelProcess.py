import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from models import Results


def formatResults(searchID):
    """
    This method populates the template excel file with information from the database
    and saves it.
    """

    wb = load_workbook('BOMOutputTemplate/BOMOutput.xlsx')
    results_data = Results.query.filter(Results.searchnumber == searchID).all()
    write = wb.active


    i=0
    totalprice = 0
    for result in results_data:

        if result.stock <= result.stockrequired or Results.stock == 0:

            write[f'C{str(8+i)}'] = result.partnumber
            write[f'D{str(8+i)}'] = result.supplier
            write[f'E{str(8+i)}'] = "Not Found"
            write[f'F{str(8+i)}'] = "Not Found"
            write[f'G{str(8+i)}'] = "Not Found"
            write[f'H{str(8+i)}'] = "Not Found"
            write[f'I{str(8+i)}'] = "Not Found"

            write[f'E{str(8 + i)}'].fill = PatternFill(patternType='solid', fgColor='00800000')

        else:
            write[f'C{str(8+i)}'] = result.partnumber
            write[f'D{str(8+i)}'] = result.supplier
            write[f'E{str(8+i)}'] = result.stock
            write[f'F{str(8+i)}'] = result.stockrequired
            write[f'G{str(8+i)}'] = result.priceperunit
            write[f'H{str(8+i)}'] = result.totalprice
            write[f'I{str(8+i)}'] = result.link

            write[f'E{str(8 + i)}'].fill = PatternFill(patternType='solid', fgColor='00008000')

            totalprice += result.totalprice

        i+=1

    write['H5'] = f'   Total price: £{round(totalprice, 2)}'
    write['H5'].font = Font(size=19)

    wb.save(f'SearchResults/BOMSearch{searchID}results.xlsx')
    wb.close()
